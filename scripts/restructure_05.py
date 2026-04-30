"""
Reestructura del INTERNO — versión unificada:
- A · Costos fijos de estructura (holding overhead)
- B · ETAPA 1 SOFTWARE Y TECNOLOGÍA — PRE-LANZAMIENTO POR MARCA (unificado)
  Columnas: Concepto | Pago único MXN | MXN anual | 8 marcas (status dropdown)
- C · ETAPA 2 — Comercialización (Marketing Q2)
- D · ETAPA 3 — Contingencia 10%
- E · Capital por invertir (Etapas 1+2+3)

+ hoja 05b · Capital histórico iBisne (mano de obra)
+ actualiza hoja 06 y Dashboard
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

PATH = r'C:/Users/ibisn/Downloads/iBisne_v5_INTERNO.xlsx'

# ============================================================================
# Estilo
# ============================================================================
NAVY = '0B1F3A'
NAVY_LIGHT = '1E3A5F'
GOLD = 'C9A961'
LIGHT_GRAY = 'F9FAFB'
MID_GRAY = 'E5E7EB'
BAND_GRAY = 'F3F4F6'
GREEN_BG = 'DCFCE7'
GREEN_LINK = '047857'
YELLOW_BG = 'FEF9C3'
RED_BG = 'FEE2E2'
BLUE_IN = '0F62FE'
YELLOW_KEY = 'FFF8B7'
FONT_NAME = 'Arial'

thin = Side(border_style='thin', color='D1D5DB')
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '"$"#,##0.00;("$"#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'
INT = '#,##0;(#,##0);"-"'

def font(size=10, bold=False, color='111827', italic=False):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color, italic=italic)
def fill(c):
    return PatternFill('solid', start_color=c, end_color=c)
def center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
def left(wrap=True, indent=0):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap, indent=indent)
def right(wrap=False):
    return Alignment(horizontal='right', vertical='center', wrap_text=wrap)
def title_cell(ws, cell, text, size=18, color=None):
    ws[cell] = text
    ws[cell].font = font(size=size, bold=True, color=color or NAVY)
    ws[cell].alignment = left(wrap=False)
def header_row(ws, row, values, start_col=1, fill_color=NAVY, text_color='FFFFFF', size=10, height=30):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col+i, value=v)
        c.font = font(size=size, bold=True, color=text_color)
        c.fill = fill(fill_color); c.alignment = center(wrap=True); c.border = BORDER_ALL
    ws.row_dimensions[row].height = height
def section_row(ws, row, text, span=10, fill_color=GOLD, text_color=NAVY, size=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font(size=size, bold=True, color=text_color)
    c.fill = fill(fill_color); c.alignment = left(wrap=False, indent=1)
    ws.row_dimensions[row].height = 26
def note_row(ws, row, text, span=10, color='6B7280'):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font(size=9, italic=True, color=color)
    c.alignment = left(wrap=True)
    ws.row_dimensions[row].height = 32

# ============================================================================
# DATOS MAESTROS
# ============================================================================
MARCAS = ['IBISNE', 'SEM ENDOMAP', 'SEM', 'MEDICAL MEXICANNA',
          'PRO FUTBOL', 'ERP ALBERCAS', 'DCI DE LA PENINSULA', 'ELIXIER']

FIJOS = [
    ('Renta local / coworking',         10000, 0,     'En curso',    'Operación Guadalajara — pago mensual'),
    ('Acta constitutiva',               0,     15000, 'Pagado',      'Pago único'),
    ('Licencia de giro comercial',      0,     675,   'Pagado',      ''),
    ('Laptop',                          0,     20000, 'Pagado',      ''),
    ('Branding e impresiones',          0,     5600,  'Pagado',      ''),
    ('Registro INDAUTOR · SEM Endomap', 0,     367,   'Pagado',      ''),
    ('Registro INDAUTOR · Pro Futbol',  0,     367,   'Pagado',      ''),
    ('Registro INDAUTOR · ERP Albercas',0,     367,   'Pagado',      ''),
]

# Unified Etapa 1 table: (concepto, tipo, pago_unico, mxn_anual, status_por_marca_lista_8)
# tipo: 'Global' = el costo se divide entre 8 marcas | 'Por marca' = costo completo cuenta por marca
# status: 'Pagado', 'Pendiente', 'En curso', 'N/A'
ETAPA1_ROWS = [
    # SOFTWARE GLOBAL (pago único cubre todos los proyectos — se prorratea 1/8 por marca)
    ('Hosting (Hostinger)',           'Global',    0,        1679.88, ['Pagado']*8),
    ('Claude 20x',                    'Global',    7400,     0,       ['Pagado']*8),
    ('Claude 5x',                     'Global',    1850,     0,       ['Pagado']*8),
    ('Google Drive',                  'Global',    0,        1990,    ['Pagado']*8),
    ('Vercel',                        'Global',    0,        4320,    ['N/A','N/A','N/A','N/A','N/A','Pagado','N/A','N/A']),
    ('Tailwind license',              'Global',    5382,     0,       ['Pagado']*8),
    # SOFTWARE POR MARCA (pago específico por proyecto — cuenta 100% a esa marca)
    ('Supabase · SEM',                'Por marca', 0,        3150,    ['N/A','N/A','Pagado','N/A','Pagado','N/A','N/A','N/A']),
    ('Supabase · Pro Futbol',         'Por marca', 0,        3150,    ['N/A','N/A','N/A','N/A','Pagado','N/A','N/A','N/A']),
    ('VPS Hetzner CX22',              'Por marca', 0,        720,     ['N/A','N/A','N/A','N/A','N/A','Pagado','N/A','N/A']),
    # LEGAL por marca
    ('Registro IMPI marca',           'Por marca', 3126.40,  0,       ['Pagado','Pendiente','Pendiente','Pagado','Pendiente','Pendiente','Pagado','Pendiente']),
    # DOMINIOS / SSL / EMAIL — cada marca con su propio
    ('Dominio principal (renovación)','Por marca', 0,        0,       ['Pendiente']*8),
    ('Dominios adicionales',          'Por marca', 0,        0,       ['Pendiente']*8),
    ('Certificado SSL premium',       'Por marca', 0,        0,       ['Pendiente']*8),
    ('Email corporativo / dominio',   'Por marca', 0,        0,       ['Pendiente']*8),
    # PLATAFORMAS / PRODUCTOS DIGITALES por marca
    ('Link bio',                      'Por marca', 0,        0,       ['Pendiente']*8),
    ('Web app',                       'Por marca', 0,        0,       ['Pendiente']*8),
    ('App Android',                   'Por marca', 0,        0,       ['Pendiente']*8),
    ('App iOS',                       'Por marca', 0,        0,       ['Pendiente']*8),
]

MANO_OBRA_TOTAL = {
    'IBISNE': 65000, 'SEM ENDOMAP': 35000, 'SEM': 200000, 'MEDICAL MEXICANNA': 35000,
    'PRO FUTBOL': 150000, 'ERP ALBERCAS': 250000, 'DCI DE LA PENINSULA': 75000, 'ELIXIER': 50000,
}
HORAS_REALES = {
    'IBISNE': 280, 'SEM ENDOMAP': 160, 'SEM': 880, 'MEDICAL MEXICANNA': 160,
    'PRO FUTBOL': 320, 'ERP ALBERCAS': 1080, 'DCI DE LA PENINSULA': 320, 'ELIXIER': 240,
}
INICIO_CONSTRUCCION = {
    'IBISNE': '01/10/2025', 'SEM ENDOMAP': '15/01/2026', 'SEM': '15/01/2026',
    'MEDICAL MEXICANNA': '15/01/2026', 'PRO FUTBOL': '01/10/2025',
    'ERP ALBERCAS': '01/10/2025', 'DCI DE LA PENINSULA': '15/01/2026', 'ELIXIER': '01/10/2025',
}
PERFILES = [('Backend', 0.30), ('Frontend', 0.25), ('MVP / Integración', 0.15),
            ('Arquitectura', 0.10), ('UX / UI', 0.10), ('QA / Pruebas', 0.10)]
TARIFA_MERCADO = 900

MKT_ACTIVAS = [0,0,1,0,1,0,1,0]  # IBISNE, SEMEN, SEM, MM, PF, ERP, DCI, ELX

# ============================================================================
# Cargar workbook y borrar hojas viejas
# ============================================================================
wb = load_workbook(PATH)
for sheet in ['05 · Inversión iBisne', '05b · Etapa 1 deploy', '05 · Recursos pendientes',
              '05b · Capital histórico']:
    if sheet in wb.sheetnames:
        del wb[sheet]

# ============================================================================
# HOJA 05 · Recursos pendientes lanzamiento
# ============================================================================
ws = wb.create_sheet('05 · Recursos pendientes')

title_cell(ws, 'A1', '05 · Recursos pendientes para lanzamiento', size=20)
ws.merge_cells('A2:K2')
ws['A2'] = ('Todo el capital que falta invertir antes y durante el lanzamiento del 04-may-2026. '
            'Mano de obra histórica en hoja "05b · Capital histórico" (no se cobra a socios nuevos).')
ws['A2'].font = font(size=10, italic=True, color='6B7280')
ws['A2'].alignment = left(wrap=True)
ws.row_dimensions[2].height = 32

# ---- A · COSTOS FIJOS ----
row = 4
section_row(ws, row, 'A · COSTOS FIJOS DE ESTRUCTURA  (organización iBisne — no se carga a marcas)', span=5); row += 1
header_row(ws, row, ['Concepto', 'MXN mensual', 'MXN anual', 'Estatus', 'Notas']); row += 1
fijos_start = row
for concepto, mens, anual, estatus, notas in FIJOS:
    ws.cell(row=row, column=1, value=concepto)
    c = ws.cell(row=row, column=2, value=mens if mens else None); c.number_format = MONEY
    c = ws.cell(row=row, column=3, value=anual if anual else None); c.number_format = MONEY
    c = ws.cell(row=row, column=4, value=estatus); c.alignment = center()
    if estatus == 'Pagado':   c.fill = fill(GREEN_BG); c.font = font(size=9, color='166534', bold=True)
    elif estatus == 'Pendiente':c.fill = fill(YELLOW_BG); c.font = font(size=9, color='854D0E', bold=True)
    else:                       c.fill = fill(BAND_GRAY); c.font = font(size=9, color='6B7280')
    ws.cell(row=row, column=5, value=notas).font = font(size=9, color='6B7280')
    for col in range(1,6): ws.cell(row=row, column=col).border = BORDER_ALL
    row += 1
fijos_end = row - 1

dv_estatus_fijos = DataValidation(type='list', formula1='"Pagado,Pendiente,En curso,Cancelado"', allow_blank=True)
dv_estatus_fijos.add(f'D{fijos_start}:D{fijos_end}')
ws.add_data_validation(dv_estatus_fijos)
ws.conditional_formatting.add(f'D{fijos_start}:D{fijos_end}',
    FormulaRule(formula=[f'$D{fijos_start}="Pagado"'], fill=fill(GREEN_BG), font=font(size=9, color='166534', bold=True)))
ws.conditional_formatting.add(f'D{fijos_start}:D{fijos_end}',
    FormulaRule(formula=[f'$D{fijos_start}="Pendiente"'], fill=fill(YELLOW_BG), font=font(size=9, color='854D0E', bold=True)))
ws.conditional_formatting.add(f'D{fijos_start}:D{fijos_end}',
    FormulaRule(formula=[f'$D{fijos_start}="En curso"'], fill=fill(BAND_GRAY), font=font(size=9, color='6B7280', bold=True)))

ws.cell(row=row, column=1, value='TOTAL COSTOS FIJOS').font = font(bold=True, color='FFFFFF')
c = ws.cell(row=row, column=2, value=f'=SUM(B{fijos_start}:B{fijos_end})'); c.number_format = MONEY; c.font = font(bold=True, color='FFFFFF')
c = ws.cell(row=row, column=3, value=f'=SUM(C{fijos_start}:C{fijos_end})'); c.number_format = MONEY; c.font = font(bold=True, color='FFFFFF')
for col in range(1,6):
    ws.cell(row=row, column=col).fill = fill(NAVY); ws.cell(row=row, column=col).border = BORDER_ALL
FIJOS_TOTAL_ROW = row   # guardamos para referenciar desde Capital por invertir
row += 3

# ============================================================================
# B · ETAPA 1 SOFTWARE Y TECNOLOGÍA — PRE-LANZAMIENTO POR MARCA (unificado)
# ============================================================================
section_row(ws, row, 'B · ETAPA 1 SOFTWARE Y TECNOLOGÍA — PRE-LANZAMIENTO POR MARCA  (unificado)', span=4+len(MARCAS)); row += 1
note_row(ws, row,
         '🎯 Columna Tipo: "Global" = el costo se divide entre 8 marcas (ej. Hosting $1,679 → $210 por marca). '
         '"Por marca" = el costo cuenta completo para cada marca que lo tiene como Pendiente. '
         'Edita las celdas AMARILLAS con los montos reales.', span=4+len(MARCAS))
row += 1

header_row(ws, row, ['Concepto', 'Tipo', 'Pago único MXN', 'MXN anual'] + MARCAS, height=32); row += 1

etapa1_data_start = row
for concepto, tipo, pago_unico, mxn_anual, statuses in ETAPA1_ROWS:
    ws.cell(row=row, column=1, value=concepto).font = font(size=10)
    # Tipo
    c = ws.cell(row=row, column=2, value=tipo)
    c.font = font(size=9, bold=True); c.alignment = center()
    if tipo == 'Global':
        c.fill = fill('DBEAFE'); c.font = font(size=9, bold=True, color='1E3A8A')
    else:
        c.fill = fill('FEE2E2'); c.font = font(size=9, bold=True, color='991B1B')
    # Pago único MXN (editable yellow)
    c = ws.cell(row=row, column=3, value=pago_unico)
    c.number_format = MONEY
    c.font = font(size=10, color=BLUE_IN, bold=True); c.fill = fill(YELLOW_KEY); c.alignment = right()
    # MXN anual (editable yellow)
    c = ws.cell(row=row, column=4, value=mxn_anual)
    c.number_format = MONEY
    c.font = font(size=10, color=BLUE_IN, bold=True); c.fill = fill(YELLOW_KEY); c.alignment = right()
    # Status per marca
    for i, stat in enumerate(statuses):
        c = ws.cell(row=row, column=5+i, value=stat)
        c.font = font(size=9, bold=True); c.alignment = center()
        if stat == 'Pagado':
            c.fill = fill(GREEN_BG); c.font = font(size=9, bold=True, color='166534')
        elif stat == 'Pendiente':
            c.fill = fill(YELLOW_BG); c.font = font(size=9, bold=True, color='854D0E')
        elif stat == 'N/A':
            c.fill = fill(BAND_GRAY); c.font = font(size=9, color='6B7280')
        elif stat == 'En curso':
            c.fill = fill('DBEAFE'); c.font = font(size=9, bold=True, color='1E3A8A')
    for col in range(1, 5+len(MARCAS)): ws.cell(row=row, column=col).border = BORDER_ALL
    row += 1
etapa1_data_end = row - 1

# Data validation: status dropdown en columnas de marca (E en adelante)
dv_status = DataValidation(type='list', formula1='"Pagado,Pendiente,En curso,N/A"', allow_blank=True)
dv_status.add(f'E{etapa1_data_start}:{get_column_letter(4+len(MARCAS))}{etapa1_data_end}')
ws.add_data_validation(dv_status)

# DV Tipo dropdown (columna B)
dv_tipo = DataValidation(type='list', formula1='"Global,Por marca"', allow_blank=True)
dv_tipo.add(f'B{etapa1_data_start}:B{etapa1_data_end}')
ws.add_data_validation(dv_tipo)

# Conditional formatting para status dinámico (cols E a E+len(MARCAS)-1)
marca_range = f'E{etapa1_data_start}:{get_column_letter(4+len(MARCAS))}{etapa1_data_end}'
ws.conditional_formatting.add(marca_range,
    FormulaRule(formula=[f'E{etapa1_data_start}="Pagado"'], fill=fill(GREEN_BG), font=font(size=9, bold=True, color='166534')))
ws.conditional_formatting.add(marca_range,
    FormulaRule(formula=[f'E{etapa1_data_start}="Pendiente"'], fill=fill(YELLOW_BG), font=font(size=9, bold=True, color='854D0E')))
ws.conditional_formatting.add(marca_range,
    FormulaRule(formula=[f'E{etapa1_data_start}="N/A"'], fill=fill(BAND_GRAY), font=font(size=9, color='6B7280')))
ws.conditional_formatting.add(marca_range,
    FormulaRule(formula=[f'E{etapa1_data_start}="En curso"'], fill=fill('DBEAFE'), font=font(size=9, bold=True, color='1E3A8A')))

# ---- TOTAL ETAPA 1 por marca: SUMPRODUCT con divisor ----
# Divisor: 8 si tipo="Global", 1 si tipo="Por marca"  →  cost / (7*(tipo="Global")+1)
ws.cell(row=row, column=1, value='TOTAL ETAPA 1 pendiente').font = font(size=11, bold=True, color='FFFFFF')
for i, m in enumerate(MARCAS):
    marca_col = get_column_letter(5+i)
    # Divisor por fila: 8 si Global, 1 si Por marca
    divisor = f'(7*($B${etapa1_data_start}:$B${etapa1_data_end}="Global")+1)'
    formula = (f'=SUMPRODUCT(({marca_col}${etapa1_data_start}:{marca_col}${etapa1_data_end}="Pendiente")'
               f'*($C${etapa1_data_start}:$C${etapa1_data_end})/{divisor})'
               f'+SUMPRODUCT(({marca_col}${etapa1_data_start}:{marca_col}${etapa1_data_end}="Pendiente")'
               f'*($D${etapa1_data_start}:$D${etapa1_data_end})/{divisor})')
    c = ws.cell(row=row, column=5+i, value=formula)
    c.number_format = MONEY; c.font = font(size=11, bold=True, color='FFFFFF'); c.alignment = right()
for col in range(1, 5+len(MARCAS)):
    ws.cell(row=row, column=col).fill = fill(NAVY); ws.cell(row=row, column=col).border = BORDER_ALL
ws.row_dimensions[row].height = 26
TOTAL_ETAPA1_ROW = row
row += 1

# Gran total Etapa 1
ws.cell(row=row, column=1, value='GRAN TOTAL ETAPA 1 — Portafolio').font = font(size=13, bold=True, color='FFFFFF')
c = ws.cell(row=row, column=5,
            value=f'=SUM({get_column_letter(5)}{TOTAL_ETAPA1_ROW}:{get_column_letter(4+len(MARCAS))}{TOTAL_ETAPA1_ROW})')
c.number_format = MONEY; c.font = font(size=13, bold=True, color='FFFFFF'); c.alignment = right()
ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=4+len(MARCAS))
for col in range(1, 5+len(MARCAS)):
    ws.cell(row=row, column=col).fill = fill('991B1B'); ws.cell(row=row, column=col).border = BORDER_ALL
ws.row_dimensions[row].height = 30
row += 3

# ============================================================================
# C · ETAPA 2 — COMERCIALIZACIÓN (Marketing Q2)
# ============================================================================
section_row(ws, row, 'C · ETAPA 2 — COMERCIALIZACIÓN  (Marketing Q2 2026 · 3 meses post-launch)', span=2+len(MARCAS)); row += 1
header_row(ws, row, ['Concepto', 'MXN mensual'] + MARCAS); row += 1
mkt_start = row
for concepto, mens in [('Contenido', 5000), ('Marketing operativo', 5000), ('Pauta comercial', 5000)]:
    ws.cell(row=row, column=1, value=concepto)
    c = ws.cell(row=row, column=2, value=mens); c.number_format = MONEY; c.font = font(color=BLUE_IN); c.fill = fill(YELLOW_KEY)
    for i, activa in enumerate(MKT_ACTIVAS):
        col = 3+i
        if activa:
            c = ws.cell(row=row, column=col, value=f'=B{row}*3'); c.number_format = MONEY
        else:
            ws.cell(row=row, column=col, value=None)
    for col in range(1, 3+len(MARCAS)): ws.cell(row=row, column=col).border = BORDER_ALL
    row += 1
mkt_end = row - 1

ws.cell(row=row, column=1, value='TOTAL ETAPA 2 — Marketing Q2').font = font(bold=True, color='FFFFFF')
for i, m in enumerate(MARCAS):
    cl = get_column_letter(3+i)
    c = ws.cell(row=row, column=3+i, value=f'=SUM({cl}{mkt_start}:{cl}{mkt_end})')
    c.number_format = MONEY; c.font = font(bold=True, color='FFFFFF')
for col in range(1, 3+len(MARCAS)):
    ws.cell(row=row, column=col).fill = fill(NAVY); ws.cell(row=row, column=col).border = BORDER_ALL
TOTAL_E2_ROW = row
row += 1

ws.cell(row=row, column=1, value='GRAN TOTAL ETAPA 2 — Portafolio').font = font(size=13, bold=True, color='FFFFFF')
c = ws.cell(row=row, column=3, value=f'=SUM({get_column_letter(3)}{TOTAL_E2_ROW}:{get_column_letter(2+len(MARCAS))}{TOTAL_E2_ROW})')
c.number_format = MONEY; c.font = font(size=13, bold=True, color='FFFFFF')
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=2+len(MARCAS))
for col in range(1, 3+len(MARCAS)):
    ws.cell(row=row, column=col).fill = fill('991B1B'); ws.cell(row=row, column=col).border = BORDER_ALL
ws.row_dimensions[row].height = 28
row += 3

# ============================================================================
# D · CAPITAL POR INVERTIR (Etapas 1 + 2 + Costos fijos A prorrateados)
# ============================================================================
section_row(ws, row, 'D · CAPITAL POR INVERTIR  (Etapa 1 + Etapa 2 + Costos fijos A prorrateados)', span=2+len(MARCAS)); row += 1
note_row(ws, row,
         '🎯 ESTE es el número que se le presenta al socio. Suma: Etapa 1 (software + deploy pendiente por marca) + '
         'Etapa 2 (marketing Q2) + Costos fijos de estructura A prorrateados /8. '
         'No incluye mano de obra histórica (capital ya invertido — ver hoja 05b · Capital histórico).',
         span=2+len(MARCAS)); row += 1
header_row(ws, row, ['Componente', 'Concepto'] + MARCAS); row += 1

# ETAPA 1
ws.cell(row=row, column=1, value='ETAPA 1').font = font(size=10, bold=True, color='854D0E')
ws.cell(row=row, column=2, value='Software + deploy pendiente por marca')
for i, m in enumerate(MARCAS):
    e1_col = get_column_letter(5+i)
    c = ws.cell(row=row, column=3+i, value=f'={e1_col}{TOTAL_ETAPA1_ROW}')
    c.number_format = MONEY; c.font = font(size=10, color=GREEN_LINK, bold=True); c.alignment = right()
for col in range(1, 3+len(MARCAS)): ws.cell(row=row, column=col).border = BORDER_ALL
PI_E1_ROW = row
row += 1

# ETAPA 2
ws.cell(row=row, column=1, value='ETAPA 2').font = font(size=10, bold=True, color='854D0E')
ws.cell(row=row, column=2, value='Comercialización (Marketing Q2)')
for i, m in enumerate(MARCAS):
    cl = get_column_letter(3+i)
    c = ws.cell(row=row, column=3+i, value=f'={cl}{TOTAL_E2_ROW}')
    c.number_format = MONEY; c.font = font(size=10, color=GREEN_LINK, bold=True); c.alignment = right()
for col in range(1, 3+len(MARCAS)): ws.cell(row=row, column=col).border = BORDER_ALL
PI_E2_ROW = row
row += 1

# COSTOS FIJOS A (prorrateado /8 por marca)
ws.cell(row=row, column=1, value='COSTOS FIJOS A').font = font(size=10, bold=True, color='854D0E')
ws.cell(row=row, column=2, value='Estructura iBisne prorrateada /8 (anual + mensual×12)')
for i, m in enumerate(MARCAS):
    # Total A annual = anual + mensual*12, dividido entre 8 marcas
    formula = f'=($C${FIJOS_TOTAL_ROW}+$B${FIJOS_TOTAL_ROW}*12)/{len(MARCAS)}'
    c = ws.cell(row=row, column=3+i, value=formula)
    c.number_format = MONEY; c.font = font(size=10, color=GREEN_LINK, bold=True); c.alignment = right()
for col in range(1, 3+len(MARCAS)): ws.cell(row=row, column=col).border = BORDER_ALL
PI_FIJOS_ROW = row
row += 1

# TOTAL POR INVERTIR
ws.cell(row=row, column=1, value='TOTAL').font = font(size=12, bold=True, color='FFFFFF')
ws.cell(row=row, column=2, value='POR INVERTIR (lo que aporta el socio)').font = font(size=12, bold=True, color='FFFFFF')
for i, m in enumerate(MARCAS):
    cl = get_column_letter(3+i)
    c = ws.cell(row=row, column=3+i, value=f'=SUM({cl}{PI_E1_ROW}:{cl}{PI_FIJOS_ROW})')
    c.number_format = MONEY; c.font = font(size=12, bold=True, color='FFFFFF'); c.alignment = right()
for col in range(1, 3+len(MARCAS)):
    ws.cell(row=row, column=col).fill = fill(NAVY); ws.cell(row=row, column=col).border = BORDER_ALL
ws.row_dimensions[row].height = 28
TOTAL_POR_INVERTIR_ROW = row
POR_INVERTIR_POR_MARCA = {m: f"'05 · Recursos pendientes'!{get_column_letter(3+i)}{row}" for i, m in enumerate(MARCAS)}
row += 2

ws.cell(row=row, column=1, value='💰 GRAN TOTAL CAPITAL POR INVERTIR — Portafolio iBisne').font = font(size=14, bold=True, color='FFFFFF')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2+len(MARCAS)-1)
c = ws.cell(row=row, column=2+len(MARCAS), value=f'=SUM({get_column_letter(3)}{TOTAL_POR_INVERTIR_ROW}:{get_column_letter(2+len(MARCAS))}{TOTAL_POR_INVERTIR_ROW})')
c.number_format = MONEY; c.font = font(size=14, bold=True, color='FFFFFF'); c.alignment = right()
for col in range(1, 3+len(MARCAS)):
    ws.cell(row=row, column=col).fill = fill('991B1B'); ws.cell(row=row, column=col).border = BORDER_ALL
ws.row_dimensions[row].height = 32
POR_INVERTIR_TOTAL_CELL = f"'05 · Recursos pendientes'!{get_column_letter(2+len(MARCAS))}{row}"
row += 3

note_row(ws, row,
         '📋  Cómo se le explica al socio:  '
         '"Tu inversión cubre 3 conceptos: (1) software y deploy específicos pendientes de tu marca, '
         '(2) marketing de los primeros 3 meses post-lanzamiento, y (3) tu parte del overhead estructural de iBisne (renta + legal + branding del holding, prorrateada). '
         'Las 3,440 horas de desarrollo ya están construidas — ver hoja 05b · Capital histórico."',
         span=2+len(MARCAS)); row += 1

# Column widths
ws.column_dimensions['A'].width = 36
ws.column_dimensions['B'].width = 12   # Tipo (Global / Por marca)
ws.column_dimensions['C'].width = 17   # Pago único MXN
ws.column_dimensions['D'].width = 17   # MXN anual
for i in range(len(MARCAS)):
    ws.column_dimensions[get_column_letter(5+i)].width = 17

ws.freeze_panes = 'E7'

# ============================================================================
# HOJA 05b · Capital histórico iBisne
# ============================================================================
ws = wb.create_sheet('05b · Capital histórico')

title_cell(ws, 'A1', '05b · Capital histórico iBisne  (capital ya invertido — no se cobra a socios)', size=18)
ws.merge_cells('A2:K2')
ws['A2'] = ('💎 Esta inversión ya fue ejecutada por iBisne. Los nuevos socios NO la pagan — la reciben construida. '
            'Se documenta para que entiendan el "valor construido" que están recibiendo.')
ws['A2'].font = font(size=10, italic=True, color='6B7280')
ws['A2'].alignment = left(wrap=True)
ws.row_dimensions[2].height = 36

row = 4
section_row(ws, row, 'A · MANO DE OBRA  (capital histórico — desglose % sobre total por marca)', span=2+len(MARCAS)); row += 1
header_row(ws, row, ['Perfil', '% default'] + MARCAS); row += 1
mo_start = row
for perfil, pct in PERFILES:
    ws.cell(row=row, column=1, value=perfil)
    c = ws.cell(row=row, column=2, value=pct); c.number_format = PCT; c.font = font(color=BLUE_IN)
    for i, m in enumerate(MARCAS):
        v = MANO_OBRA_TOTAL[m] * pct
        c = ws.cell(row=row, column=3+i, value=v); c.number_format = MONEY
    for col in range(1, 3+len(MARCAS)): ws.cell(row=row, column=col).border = BORDER_ALL
    row += 1
mo_end = row - 1

ws.cell(row=row, column=1, value='TOTAL MANO DE OBRA (capital histórico)').font = font(bold=True, color='FFFFFF')
ws.cell(row=row, column=2, value=f'=SUM(B{mo_start}:B{mo_end})').number_format = PCT
for i, m in enumerate(MARCAS):
    cl = get_column_letter(3+i)
    c = ws.cell(row=row, column=3+i, value=f'=SUM({cl}{mo_start}:{cl}{mo_end})')
    c.number_format = MONEY; c.font = font(bold=True, color='FFFFFF')
for col in range(1, 3+len(MARCAS)):
    ws.cell(row=row, column=col).fill = fill(NAVY); ws.cell(row=row, column=col).border = BORDER_ALL
mo_totals_row = row
row += 1

ws.cell(row=row, column=1, value='Horas reales invertidas').font = font(size=9, bold=True, color=NAVY)
for i, m in enumerate(MARCAS):
    c = ws.cell(row=row, column=3+i, value=HORAS_REALES[m])
    c.number_format = INT; c.font = font(size=9, color=BLUE_IN, bold=True); c.fill = fill(YELLOW_KEY); c.alignment = center()
for col in range(1, 3+len(MARCAS)): ws.cell(row=row, column=col).border = BORDER_ALL
horas_row = row
row += 1

ws.cell(row=row, column=1, value='Tarifa interna iBisne (calc)').font = font(size=9, italic=True, color='6B7280')
for i, m in enumerate(MARCAS):
    cl = get_column_letter(3+i)
    c = ws.cell(row=row, column=3+i, value=f'=IFERROR({cl}{mo_totals_row}/{cl}{horas_row},0)')
    c.number_format = '"$"#,##0"/hr"'; c.font = font(size=9, italic=True, color='6B7280'); c.alignment = center()
for col in range(1, 3+len(MARCAS)): ws.cell(row=row, column=col).border = BORDER_ALL
row += 1

ws.cell(row=row, column=1, value=f'Valor de mercado equivalente  (× ${TARIFA_MERCADO}/hr)').font = font(size=10, bold=True, color=GREEN_LINK)
for i, m in enumerate(MARCAS):
    cl = get_column_letter(3+i)
    c = ws.cell(row=row, column=3+i, value=f'={cl}{horas_row}*{TARIFA_MERCADO}')
    c.number_format = MONEY; c.font = font(size=10, bold=True, color=GREEN_LINK); c.fill = fill(GREEN_BG); c.alignment = right()
for col in range(1, 3+len(MARCAS)): ws.cell(row=row, column=col).border = BORDER_ALL
valor_mercado_row = row
row += 1

ws.cell(row=row, column=1, value='Inicio de construcción').font = font(size=9, italic=True, color='6B7280')
for i, m in enumerate(MARCAS):
    c = ws.cell(row=row, column=3+i, value=INICIO_CONSTRUCCION[m])
    c.font = font(size=9, italic=True, color='6B7280'); c.alignment = center()
for col in range(1, 3+len(MARCAS)): ws.cell(row=row, column=col).border = BORDER_ALL
row += 2

ws.cell(row=row, column=1, value='GRAN TOTAL — Costo real iBisne').font = font(size=11, bold=True, color='FFFFFF')
c = ws.cell(row=row, column=3, value=f'=SUM({get_column_letter(3)}{mo_totals_row}:{get_column_letter(2+len(MARCAS))}{mo_totals_row})')
c.number_format = MONEY; c.font = font(bold=True, color='FFFFFF')
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=2+len(MARCAS))
for col in range(1, 3+len(MARCAS)):
    ws.cell(row=row, column=col).fill = fill(NAVY_LIGHT); ws.cell(row=row, column=col).border = BORDER_ALL
row += 1

ws.cell(row=row, column=1, value='GRAN TOTAL — Horas reales invertidas').font = font(size=10, bold=True, color='FFFFFF')
c = ws.cell(row=row, column=3, value=f'=SUM({get_column_letter(3)}{horas_row}:{get_column_letter(2+len(MARCAS))}{horas_row})')
c.number_format = INT + ' "hrs"'; c.font = font(bold=True, color='FFFFFF')
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=2+len(MARCAS))
for col in range(1, 3+len(MARCAS)):
    ws.cell(row=row, column=col).fill = fill(NAVY); ws.cell(row=row, column=col).border = BORDER_ALL
row += 1

ws.cell(row=row, column=1, value='GRAN TOTAL — Valor de mercado equivalente').font = font(size=12, bold=True, color='FFFFFF')
c = ws.cell(row=row, column=3, value=f'=SUM({get_column_letter(3)}{valor_mercado_row}:{get_column_letter(2+len(MARCAS))}{valor_mercado_row})')
c.number_format = MONEY; c.font = font(size=12, bold=True, color='FFFFFF')
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=2+len(MARCAS))
for col in range(1, 3+len(MARCAS)):
    ws.cell(row=row, column=col).fill = fill('10B981'); ws.cell(row=row, column=col).border = BORDER_ALL
ws.row_dimensions[row].height = 28
row += 2

note_row(ws, row,
         '💡  iBisne entregó 3,440 horas de equipo experto que a tarifas estándar MX 2026 valdrían ~$3,096,000 MXN. '
         'Costo real iBisne: $860,000 → tarifa interna ~$250/hr vs. $900/hr de mercado = 3.6× de eficiencia.',
         span=2+len(MARCAS))
row += 3

# Capital histórico por marca
section_row(ws, row, 'B · CAPITAL HISTÓRICO TOTAL POR MARCA', span=2+len(MARCAS)); row += 1
header_row(ws, row, ['Componente', 'Concepto'] + MARCAS); row += 1
hist_block_start = row

ws.cell(row=row, column=1, value='YA INVERTIDO').font = font(size=9, bold=True, color=GREEN_LINK)
ws.cell(row=row, column=2, value='Mano de obra histórica')
for i, m in enumerate(MARCAS):
    cl = get_column_letter(3+i)
    c = ws.cell(row=row, column=3+i, value=f'={cl}{mo_totals_row}')
    c.number_format = MONEY; c.font = font(color=GREEN_LINK)
for col in range(1, 3+len(MARCAS)): ws.cell(row=row, column=col).border = BORDER_ALL
row += 1

# SUBTOTAL CAPITAL HISTÓRICO (solo mano de obra porque software y fijos ya los vemos en hoja 05 separado)
ws.cell(row=row, column=1, value='SUBTOTAL').font = font(size=12, bold=True, color='FFFFFF')
ws.cell(row=row, column=2, value='CAPITAL HISTÓRICO POR MARCA').font = font(size=12, bold=True, color='FFFFFF')
for i, m in enumerate(MARCAS):
    cl = get_column_letter(3+i)
    c = ws.cell(row=row, column=3+i, value=f'=SUM({cl}{hist_block_start}:{cl}{row-1})')
    c.number_format = MONEY; c.font = font(size=11, bold=True, color='FFFFFF')
for col in range(1, 3+len(MARCAS)):
    ws.cell(row=row, column=col).fill = fill(NAVY); ws.cell(row=row, column=col).border = BORDER_ALL
ws.row_dimensions[row].height = 28
HIST_TOTAL_ROW = row
HIST_POR_MARCA = {m: f"'05b · Capital histórico'!{get_column_letter(3+i)}{row}" for i, m in enumerate(MARCAS)}

row += 2
ws.cell(row=row, column=1, value='💎 GRAN TOTAL CAPITAL HISTÓRICO — Portafolio iBisne').font = font(size=14, bold=True, color='FFFFFF')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2+len(MARCAS)-1)
c = ws.cell(row=row, column=2+len(MARCAS), value=f'=SUM({get_column_letter(3)}{HIST_TOTAL_ROW}:{get_column_letter(2+len(MARCAS))}{HIST_TOTAL_ROW})')
c.number_format = MONEY; c.font = font(size=14, bold=True, color='FFFFFF'); c.alignment = right()
for col in range(1, 3+len(MARCAS)):
    ws.cell(row=row, column=col).fill = fill(NAVY_LIGHT); ws.cell(row=row, column=col).border = BORDER_ALL
ws.row_dimensions[row].height = 32
HIST_TOTAL_CELL = f"'05b · Capital histórico'!{get_column_letter(2+len(MARCAS))}{row}"

ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 38
for i in range(len(MARCAS)):
    ws.column_dimensions[get_column_letter(3+i)].width = 17
ws.freeze_panes = 'C4'

# ============================================================================
# Hoja 06 P&L: actualizar refs a Capital histórico + Capital por invertir
# ============================================================================
ws6 = wb['06 · P&L proyectado']

for i, m in enumerate(MARCAS):
    r = 6 + i
    formula_be = f'=IFERROR(({HIST_POR_MARCA[m]}+{POR_INVERTIR_POR_MARCA[m]})/(F{r}*H{r}),"N/A")'
    c = ws6.cell(row=r, column=12, value=formula_be)
    c.number_format = '0.0" meses";"N/A";"-"'; c.alignment = center(); c.font = font(bold=True)
    formula_roi = f'=IFERROR((K{r}-({HIST_POR_MARCA[m]}+{POR_INVERTIR_POR_MARCA[m]}))/({HIST_POR_MARCA[m]}+{POR_INVERTIR_POR_MARCA[m]}),0)'
    c = ws6.cell(row=r, column=13, value=formula_roi)
    c.number_format = PCT; c.font = font(bold=True)

# Total portafolio row 14
c = ws6.cell(row=14, column=13, value=f'=IFERROR((K14-({HIST_TOTAL_CELL}+{POR_INVERTIR_TOTAL_CELL}))/({HIST_TOTAL_CELL}+{POR_INVERTIR_TOTAL_CELL}),0)')
c.number_format = PCT; c.font = font(bold=True, color='FFFFFF')

# Tabla doble perspectiva (buscar dinámicamente)
for r in range(15, 50):
    val = ws6.cell(r, 1).value
    if val == 'IBISNE':
        for i, m in enumerate(MARCAS):
            rr = r + i
            ws6.cell(row=rr, column=2, value=f'={HIST_POR_MARCA[m]}').number_format = MONEY
            ws6.cell(row=rr, column=2).font = font(color=GREEN_LINK, size=9)
            ws6.cell(row=rr, column=3, value=f'={POR_INVERTIR_POR_MARCA[m]}').number_format = MONEY
            ws6.cell(row=rr, column=3).font = font(color='854D0E', bold=True, size=9)
            ws6.cell(row=rr, column=4, value=f'=B{rr}+C{rr}').number_format = MONEY
            ws6.cell(row=rr, column=4).font = font(bold=True, size=9)
            ws6.cell(row=rr, column=5, value=f'=K{6+i}/12').number_format = MONEY
            ws6.cell(row=rr, column=5).font = font(size=9)
            ws6.cell(row=rr, column=6, value=f'=IFERROR(D{rr}/E{rr},"N/A")')
            ws6.cell(row=rr, column=6).number_format = '0.0" meses";"N/A";"-"'
            ws6.cell(row=rr, column=6).alignment = center()
            ws6.cell(row=rr, column=6).font = font(bold=True, size=9, color=NAVY)
            ws6.cell(row=rr, column=7, value=f'=IFERROR(C{rr}/E{rr},"N/A")')
            ws6.cell(row=rr, column=7).number_format = '0.0" meses";"N/A";"-"'
            ws6.cell(row=rr, column=7).alignment = center()
            ws6.cell(row=rr, column=7).font = font(bold=True, size=9, color=GREEN_LINK)
            ws6.cell(row=rr, column=8, value=f'=IFERROR(F{rr}-G{rr},"N/A")')
            ws6.cell(row=rr, column=8).number_format = '0.0" meses";"N/A";"-"'
            ws6.cell(row=rr, column=8).alignment = center()
            ws6.cell(row=rr, column=8).font = font(size=9, italic=True, color='6B7280')
        total_r = r + 8
        ws6.cell(row=total_r, column=2, value=f'={HIST_TOTAL_CELL}').number_format = MONEY
        ws6.cell(row=total_r, column=2).font = font(bold=True, color='FFFFFF', size=9)
        ws6.cell(row=total_r, column=3, value=f'={POR_INVERTIR_TOTAL_CELL}').number_format = MONEY
        ws6.cell(row=total_r, column=3).font = font(bold=True, color='FFFFFF', size=9)
        ws6.cell(row=total_r, column=4, value=f'=B{total_r}+C{total_r}').number_format = MONEY
        ws6.cell(row=total_r, column=4).font = font(bold=True, color='FFFFFF', size=9)
        ws6.cell(row=total_r, column=5, value=f'=SUM(E{r}:E{r+7})').number_format = MONEY
        ws6.cell(row=total_r, column=5).font = font(bold=True, color='FFFFFF', size=9)
        ws6.cell(row=total_r, column=6, value=f'=IFERROR(D{total_r}/E{total_r},"N/A")')
        ws6.cell(row=total_r, column=6).number_format = '0.0" meses"'
        ws6.cell(row=total_r, column=6).alignment = center()
        ws6.cell(row=total_r, column=6).font = font(bold=True, color='FFFFFF', size=9)
        ws6.cell(row=total_r, column=7, value=f'=IFERROR(C{total_r}/E{total_r},"N/A")')
        ws6.cell(row=total_r, column=7).number_format = '0.0" meses"'
        ws6.cell(row=total_r, column=7).alignment = center()
        ws6.cell(row=total_r, column=7).font = font(bold=True, color='FFFFFF', size=9)
        ws6.cell(row=total_r, column=8, value=f'=IFERROR(F{total_r}-G{total_r},"N/A")')
        ws6.cell(row=total_r, column=8).number_format = '0.0" meses"'
        ws6.cell(row=total_r, column=8).alignment = center()
        ws6.cell(row=total_r, column=8).font = font(bold=True, color='FFFFFF', size=9)
        break

# ============================================================================
# Dashboard
# ============================================================================
ws0 = wb['00 · Dashboard']
ws0.cell(row=15, column=1, value=f'={HIST_TOTAL_CELL}').number_format = MONEY
ws0.cell(row=15, column=5, value=f'={POR_INVERTIR_TOTAL_CELL}').number_format = MONEY
ws0.cell(row=15, column=9, value=f'={HIST_TOTAL_CELL}+{POR_INVERTIR_TOTAL_CELL}').number_format = MONEY
for i, m in enumerate(MARCAS):
    r = 25 + i
    ws0.cell(row=r, column=5, value=f'={HIST_POR_MARCA[m]}').number_format = MONEY
    ws0.cell(row=r, column=5).font = font(color=GREEN_LINK, size=9)
    ws0.cell(row=r, column=6, value=f'={POR_INVERTIR_POR_MARCA[m]}').number_format = MONEY
    ws0.cell(row=r, column=6).font = font(color='854D0E', bold=True, size=9)
ws0.cell(row=33, column=5, value=f'={HIST_TOTAL_CELL}').number_format = MONEY
ws0.cell(row=33, column=5).font = font(bold=True, color='FFFFFF', size=9)
ws0.cell(row=33, column=6, value=f'={POR_INVERTIR_TOTAL_CELL}').number_format = MONEY
ws0.cell(row=33, column=6).font = font(bold=True, color='FFFFFF', size=9)

# ============================================================================
# Reordenar hojas
# ============================================================================
order = ['00 · Dashboard', '01 · Portafolio', '02 · Avance', '03 · Canales',
         '04 · Reglas y plazos', '05 · Recursos pendientes', '05b · Capital histórico',
         '06 · P&L proyectado', '07 · Glosario', '08 · Bitácora']
wb._sheets = [wb[n] for n in order]
wb.active = 0

wb.save(PATH)
print(f'OK -> {PATH}')
print(f'Etapa 1 unified table: 22 rows × 8 marcas columnas con status dropdown')
