"""
Actualiza el modelo P&L de los 2 archivos v5:
  · Costo unitario = producto BASE puro (1 gomita = $50, no blended)
  · Vol base y optimista como fórmulas: conservador × multiplicador
  · Multiplicadores globales editables: base × 3, optimista × 10

Edita las celdas amarillas y todo recalcula. Si quieres override manual una marca,
escribes el valor sobre la fórmula.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

NAVY = '0B1F3A'
BLUE_IN = '0F62FE'
GREEN_LINK = '047857'
YELLOW_KEY = 'FFF8B7'

def font(size=10, bold=False, color='111827', italic=False):
    return Font(name='Arial', size=size, bold=bold, color=color, italic=italic)
def fill(c):
    return PatternFill('solid', start_color=c, end_color=c)

# Datos actualizados (producto base puro + vol conservador)
DATA = {
    'IBISNE':              {'precio': 50000, 'costo': 15000, 'cons': 1},
    'SEM ENDOMAP':         {'precio': 500,   'costo': 150,   'cons': 20},
    'SEM':                 {'precio': 120,   'costo': 50,    'cons': 500},
    'MEDICAL MEXICANNA':   {'precio': 120,   'costo': 50,    'cons': 200},
    'PRO FUTBOL':          {'precio': 25,    'costo': 0,     'cons': 500},
    'ERP ALBERCAS':        {'precio': 947,   'costo': 100,   'cons': 5},
    'DCI DE LA PENINSULA': {'precio': 50000, 'costo': 5000,  'cons': 1},
    'ELIXIER':             {'precio': 197,   'costo': 35,    'cons': 50},
}

def update_internal(path):
    wb = load_workbook(path)
    ws = wb['06 · P&L proyectado']
    # Escribir multiplicadores en R3 (libre)
    ws['A3'] = '🎚️ Multiplicadores escenario:'
    ws['A3'].font = font(size=10, bold=True, color=NAVY)
    ws['B3'] = 'Base ='
    ws['B3'].font = font(size=10, bold=True, color='6B7280')
    ws['B3'].alignment = Alignment(horizontal='right', vertical='center')
    ws['C3'] = 3
    ws['C3'].font = font(size=11, bold=True, color=BLUE_IN)
    ws['C3'].fill = fill(YELLOW_KEY)
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C3'].number_format = '0"× conservador"'
    ws['D3'] = 'Optimista ='
    ws['D3'].font = font(size=10, bold=True, color='6B7280')
    ws['D3'].alignment = Alignment(horizontal='right', vertical='center')
    ws['E3'] = 10
    ws['E3'].font = font(size=11, bold=True, color=BLUE_IN)
    ws['E3'].fill = fill(YELLOW_KEY)
    ws['E3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E3'].number_format = '0"× conservador"'
    ws['F3'] = '← edita estos para escalar todos los escenarios a la vez'
    ws['F3'].font = font(size=9, italic=True, color='6B7280')
    ws['F3'].alignment = Alignment(horizontal='left', vertical='center')

    # Filas de marcas: 6 a 13. Vol conservador G, base H, optimista I.
    marcas_rows = {
        'IBISNE': 6, 'SEM ENDOMAP': 7, 'SEM': 8, 'MEDICAL MEXICANNA': 9,
        'PRO FUTBOL': 10, 'ERP ALBERCAS': 11, 'DCI DE LA PENINSULA': 12, 'ELIXIER': 13,
    }
    money_fmt = '"$"#,##0.00;("$"#,##0.00);"-"'
    for marca, r in marcas_rows.items():
        d = DATA[marca]
        # Precio y costo (D y E)
        c = ws.cell(row=r, column=4, value=d['precio'])
        c.number_format = money_fmt
        c.font = font(color=BLUE_IN, bold=True)
        c.fill = fill(YELLOW_KEY)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c = ws.cell(row=r, column=5, value=d['costo'])
        c.number_format = money_fmt
        c.font = font(color=BLUE_IN)
        c.fill = fill(YELLOW_KEY)
        c.alignment = Alignment(horizontal='right', vertical='center')
        # Vol conservador (G)
        c = ws.cell(row=r, column=7, value=d['cons'])
        c.font = font(color=BLUE_IN, bold=True)
        c.fill = fill(YELLOW_KEY)
        c.alignment = Alignment(horizontal='center', vertical='center')
        # Vol base (H) = conservador × $C$3
        c = ws.cell(row=r, column=8, value=f'=G{r}*$C$3')
        c.font = font(color=GREEN_LINK, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
        # Vol optimista (I) = conservador × $E$3
        c = ws.cell(row=r, column=9, value=f'=G{r}*$E$3')
        c.font = font(color=GREEN_LINK, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(path)
    print(f'OK INTERNO -> {path}')

def update_pitch(path):
    wb = load_workbook(path)
    ws = wb['03 · Proyección 3 escenarios']
    # Multiplicadores en R3
    ws['A3'] = '🎚️ Multiplicadores escenario:'
    ws['A3'].font = font(size=10, bold=True, color=NAVY)
    ws['B3'] = 'Base ='
    ws['B3'].font = font(size=10, bold=True, color='6B7280')
    ws['B3'].alignment = Alignment(horizontal='right', vertical='center')
    ws['C3'] = 3
    ws['C3'].font = font(size=11, bold=True, color=BLUE_IN)
    ws['C3'].fill = fill(YELLOW_KEY)
    ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C3'].number_format = '0"× conservador"'
    ws['D3'] = 'Optimista ='
    ws['D3'].font = font(size=10, bold=True, color='6B7280')
    ws['D3'].alignment = Alignment(horizontal='right', vertical='center')
    ws['E3'] = 10
    ws['E3'].font = font(size=11, bold=True, color=BLUE_IN)
    ws['E3'].fill = fill(YELLOW_KEY)
    ws['E3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['E3'].number_format = '0"× conservador"'
    ws['F3'] = '← edita para escalar todos los escenarios'
    ws['F3'].font = font(size=9, italic=True, color='6B7280')
    ws['F3'].alignment = Alignment(horizontal='left', vertical='center')

    # PITCH: Vol conservador F, base G, optimista H. Marcas filas 5-12.
    marcas_rows = {
        'IBISNE': 5, 'SEM ENDOMAP': 6, 'SEM': 7, 'MEDICAL MEXICANNA': 8,
        'PRO FUTBOL': 9, 'ERP ALBERCAS': 10, 'DCI DE LA PENINSULA': 11, 'ELIXIER': 12,
    }
    money_fmt = '"$"#,##0.00;("$"#,##0.00);"-"'
    for marca, r in marcas_rows.items():
        d = DATA[marca]
        # Precio C, costo D
        c = ws.cell(row=r, column=3, value=d['precio'])
        c.number_format = money_fmt; c.font = font(color=BLUE_IN, bold=True); c.fill = fill(YELLOW_KEY)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c = ws.cell(row=r, column=4, value=d['costo'])
        c.number_format = money_fmt; c.font = font(color=BLUE_IN); c.fill = fill(YELLOW_KEY)
        c.alignment = Alignment(horizontal='right', vertical='center')
        # Vol conservador F
        c = ws.cell(row=r, column=6, value=d['cons'])
        c.font = font(color=BLUE_IN, bold=True); c.fill = fill(YELLOW_KEY)
        c.alignment = Alignment(horizontal='center', vertical='center')
        # Vol base G = F × $C$3
        c = ws.cell(row=r, column=7, value=f'=F{r}*$C$3')
        c.font = font(color=GREEN_LINK, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
        # Vol optimista H = F × $E$3
        c = ws.cell(row=r, column=8, value=f'=F{r}*$E$3')
        c.font = font(color=GREEN_LINK, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(path)
    print(f'OK PITCH -> {path}')

if __name__ == '__main__':
    update_internal(r'C:/Users/ibisn/Downloads/iBisne_v5_INTERNO.xlsx')
    update_pitch(r'C:/Users/ibisn/Downloads/iBisne_v5_Inversionistas.xlsx')
