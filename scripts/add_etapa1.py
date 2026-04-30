"""
Agrega al INTERNO la hoja "05b · Etapa 1 deploy" con 10 conceptos por marca.
Reorganiza el bloque POR INVERTIR de la hoja 05 a la estructura:
  Etapa 1 (pre-lanzamiento, deploy)  ←  pull de hoja 05b
  Etapa 2 (comercialización Q2)
  Etapa 3 (contingencia 10%)

La renta sale de "Por invertir" (la absorbe el holding como overhead).
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

PATH = r'C:/Users/ibisn/Downloads/iBisne_v5_INTERNO.xlsx'

# ============================================================================
# Estilo
# ============================================================================
NAVY = '0B1F3A'
NAVY_LIGHT = '1E3A5F'
GOLD = 'C9A961'
LIGHT_GRAY = 'F9FAFB'
MID_GRAY = 'E5E7EB'
BAND_GRAY = 'F3F4F6'
GREEN = '10B981'
GREEN_BG = 'DCFCE7'
GREEN_LINK = '047857'
YELLOW_BG = 'FEF9C3'
RED_BG = 'FEE2E2'
BLUE_IN = '0F62FE'
YELLOW_KEY = 'FFF8B7'
FONT_NAME = 'Arial'

thin = Side(border_style='thin', color='D1D5DB')
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '"$"#,##0.00;("$"#,##0.00);"-"'

def font(size=10, bold=False, color='111827', italic=False):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color, italic=italic)
def fill(c):
    return PatternFill('solid', start_color=c, end_color=c)
def center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
def left(wrap=True, indent=0):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap, indent=indent)
def right(wrap=False):
    return Alignment(horizontal='right', vertical='center', wrap_text=wrap)

# ============================================================================
# Datos: 8 marcas × 10 conceptos
# ============================================================================
MARCAS = ['IBISNE', 'SEM ENDOMAP', 'SEM', 'MEDICAL MEXICANNA',
          'PRO FUTBOL', 'ERP ALBERCAS', 'DCI DE LA PENINSULA', 'ELIXIER']

# Marcas con IMPI pendiente (✅ = ya tienen marca registrada)
IMPI_PENDIENTE = {
    'IBISNE': False,
    'SEM ENDOMAP': True,
    'SEM': True,
    'MEDICAL MEXICANNA': False,
    'PRO FUTBOL': True,
    'ERP ALBERCAS': True,
    'DCI DE LA PENINSULA': False,
    'ELIXIER': True,
}

# 10 conceptos típicos de pre-lanzamiento por marca
# (concepto, monto pre-llenado o None, estatus default, nota)
CONCEPTOS_TIPICOS = [
    ('Registro IMPI marca',                 'IMPI',       None),  # se calcula según marca
    ('Hosting / Vercel pro plan',           None,         'Pendiente', 'Plan que aguante tráfico de lanzamiento'),
    ('Renovación dominio principal',        None,         'Pendiente', 'Vence/se compra .com.mx'),
    ('Dominios adicionales',                None,         'Pendiente', 'Variantes .com / .mx / redirects'),
    ('Certificado SSL premium',             None,         'Pendiente', 'HTTPS sin warnings, sello de confianza'),
    ('Setup pasarela de pago',              None,         'Pendiente', 'Configuración MercadoPago / OpenPay / Stripe'),
    ('Configuración CRM inicial',           None,         'Pendiente', 'Plantillas, secuencias, integraciones'),
    ('Email corporativo / dominio',         None,         'Pendiente', 'Cuentas Google Workspace / Zoho'),
    ('Fotografía de producto / branding',   None,         'Pendiente', 'Set de fotos para sitio + redes'),
    ('Servicios técnicos pendientes',       None,         'Pendiente', 'Fixes, testing QA, integraciones extra'),
]

# ============================================================================
# Cargar workbook
# ============================================================================
wb = load_workbook(PATH)

# Si la hoja 05b ya existe (por re-correr), la borramos
if '05b · Etapa 1 deploy' in wb.sheetnames:
    del wb['05b · Etapa 1 deploy']

# ============================================================================
# Crear hoja 05b · Etapa 1 deploy
# ============================================================================
ws = wb.create_sheet('05b · Etapa 1 deploy')
ws.sheet_view.showGridLines = False

# Título
ws['A1'] = 'ETAPA 1 — PRE-LANZAMIENTO POR MARCA  (deploy)'
ws['A1'].font = font(size=20, bold=True, color=NAVY)
ws['A1'].alignment = left(wrap=False)
ws.row_dimensions[1].height = 30

ws.merge_cells('A2:F2')
ws['A2'] = ('Lo que falta pagar para que cada proyecto pueda salir live el 04-may-2026. '
            'Edita las celdas AMARILLAS con los montos reales. Los totales por marca se reflejan automáticamente en la hoja 05 (bloque POR INVERTIR).')
ws['A2'].font = font(size=10, italic=True, color='6B7280')
ws['A2'].alignment = left(wrap=True)
ws.row_dimensions[2].height = 32

ws.merge_cells('A3:F3')
ws['A3'] = ('💡  Cada marca tiene 10 conceptos típicos. Si una marca no necesita algún concepto, déjalo en $0 (Estatus: "N/A"). '
            'Si te falta un concepto que no aparece, agrégalo en la columna "Notas" y suma manualmente al monto.')
ws['A3'].font = font(size=9, italic=True, color='6B7280')
ws['A3'].alignment = left(wrap=True)
ws.row_dimensions[3].height = 28

# Headers tabla
HEADER_ROW = 5
headers = ['Marca', '#', 'Concepto', 'Monto (MXN)', 'Estatus', 'Notas']
for i, h in enumerate(headers):
    c = ws.cell(row=HEADER_ROW, column=i+1, value=h)
    c.font = font(size=10, bold=True, color='FFFFFF')
    c.fill = fill(NAVY); c.alignment = center(wrap=True); c.border = BORDER_ALL
ws.row_dimensions[HEADER_ROW].height = 28

# Body — por cada marca: 10 conceptos + 1 total
row = HEADER_ROW + 1
totals_per_marca = {}
estatus_dv_ranges = []

for m_idx, marca in enumerate(MARCAS):
    marca_start = row
    # Banner de marca (toda la fila)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value=f'  📌  {marca}')
    c.font = font(size=12, bold=True, color='FFFFFF')
    c.fill = fill(NAVY_LIGHT); c.alignment = left(wrap=False, indent=1)
    ws.row_dimensions[row].height = 22
    row += 1

    items_start = row
    for i, (concepto, mark, *rest) in enumerate(CONCEPTOS_TIPICOS):
        # Marca: solo en la primera fila de cada bloque (visual: vacío en las demás)
        c = ws.cell(row=row, column=1, value=marca if i == 0 else '')
        c.font = font(size=9, color='6B7280')
        c.alignment = left()
        c.border = BORDER_ALL
        # #
        c = ws.cell(row=row, column=2, value=i+1)
        c.font = font(size=9, color='6B7280'); c.alignment = center(); c.border = BORDER_ALL
        # Concepto
        c = ws.cell(row=row, column=3, value=concepto)
        c.font = font(size=10); c.alignment = left(); c.border = BORDER_ALL

        # Monto: IMPI pre-lleno, demás vacío
        if mark == 'IMPI':
            valor = 3126.40 if IMPI_PENDIENTE[marca] else 0
            estatus = 'Pendiente' if IMPI_PENDIENTE[marca] else 'Pagado'
            nota = 'Tarifa oficial IMPI México 2026' if IMPI_PENDIENTE[marca] else 'Marca ya registrada'
        else:
            valor = 0  # vacío para que el user llene
            estatus = rest[0] if rest else 'Pendiente'
            nota = rest[1] if len(rest) > 1 else ''

        c = ws.cell(row=row, column=4, value=valor)
        c.number_format = MONEY
        c.font = font(size=10, color=BLUE_IN, bold=True)
        c.fill = fill(YELLOW_KEY)
        c.alignment = right(); c.border = BORDER_ALL
        # Estatus
        c = ws.cell(row=row, column=5, value=estatus)
        c.font = font(size=9, bold=True)
        c.alignment = center(); c.border = BORDER_ALL
        # Color según estatus
        if estatus == 'Pagado':
            c.fill = fill(GREEN_BG); c.font = font(size=9, bold=True, color='166534')
        elif estatus == 'Pendiente':
            c.fill = fill(YELLOW_BG); c.font = font(size=9, bold=True, color='854D0E')
        elif estatus == 'N/A':
            c.fill = fill(BAND_GRAY); c.font = font(size=9, color='6B7280')
        # Notas
        c = ws.cell(row=row, column=6, value=nota)
        c.font = font(size=9, italic=True, color='6B7280')
        c.alignment = left(); c.border = BORDER_ALL
        row += 1

    items_end = row - 1
    # Total por marca
    c = ws.cell(row=row, column=1, value='')
    c.border = BORDER_ALL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value=f'TOTAL ETAPA 1 · {marca}')
    c.font = font(size=11, bold=True, color='FFFFFF')
    c.fill = fill(NAVY); c.alignment = right(); c.border = BORDER_ALL
    # Total formula
    c = ws.cell(row=row, column=4, value=f'=SUM(D{items_start}:D{items_end})')
    c.number_format = MONEY
    c.font = font(size=12, bold=True, color='FFFFFF')
    c.fill = fill(NAVY); c.alignment = right(); c.border = BORDER_ALL
    ws.cell(row=row, column=5).fill = fill(NAVY); ws.cell(row=row, column=5).border = BORDER_ALL
    ws.cell(row=row, column=6).fill = fill(NAVY); ws.cell(row=row, column=6).border = BORDER_ALL
    ws.row_dimensions[row].height = 24
    totals_per_marca[marca] = row  # guardamos qué fila tiene el total de cada marca
    row += 1

    # Espaciado entre marcas
    row += 1

# Estatus dropdown (data validation)
dv = DataValidation(type='list', formula1='"Pagado,Pendiente,En curso,N/A"', allow_blank=True)
# Aplicar a TODA la columna E (filas 7 hasta el final), excepto totales
# Para simplicidad, lo aplico a un rango grande
dv.add(f'E6:E{row}')
ws.add_data_validation(dv)

# Conditional formatting para Estatus dinámico
for r in range(6, row):
    if r in totals_per_marca.values():
        continue
ws.conditional_formatting.add(f'E6:E{row}',
    FormulaRule(formula=[f'$E6="Pagado"'], fill=fill(GREEN_BG), font=font(size=9, bold=True, color='166534')))
ws.conditional_formatting.add(f'E6:E{row}',
    FormulaRule(formula=[f'$E6="Pendiente"'], fill=fill(YELLOW_BG), font=font(size=9, bold=True, color='854D0E')))
ws.conditional_formatting.add(f'E6:E{row}',
    FormulaRule(formula=[f'$E6="N/A"'], fill=fill(BAND_GRAY), font=font(size=9, color='6B7280')))
ws.conditional_formatting.add(f'E6:E{row}',
    FormulaRule(formula=[f'$E6="En curso"'], fill=fill('DBEAFE'), font=font(size=9, bold=True, color='1E3A8A')))

# Gran total Etapa 1 portafolio
row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
c = ws.cell(row=row, column=1, value='GRAN TOTAL ETAPA 1 — Portafolio')
c.font = font(size=14, bold=True, color='FFFFFF')
c.fill = fill('991B1B'); c.alignment = right(); c.border = BORDER_ALL
# Sum de todos los totales por marca
sum_formula = '+'.join([f'D{totals_per_marca[m]}' for m in MARCAS])
c = ws.cell(row=row, column=4, value=f'={sum_formula}')
c.number_format = MONEY
c.font = font(size=14, bold=True, color='FFFFFF')
c.fill = fill('991B1B'); c.alignment = right(); c.border = BORDER_ALL
ws.cell(row=row, column=5).fill = fill('991B1B'); ws.cell(row=row, column=5).border = BORDER_ALL
ws.cell(row=row, column=6).fill = fill('991B1B'); ws.cell(row=row, column=6).border = BORDER_ALL
ws.row_dimensions[row].height = 28
GRAN_TOTAL_E1_ROW = row

# Column widths
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 4
ws.column_dimensions['C'].width = 38
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 50

# Freeze panes
ws.freeze_panes = 'A6'

# ============================================================================
# Modificar hoja 05 · bloque POR INVERTIR
# ============================================================================
ws05 = wb['05 · Inversión iBisne']

# Row 73: cambiar IMPI → Etapa 1 (pull de 05b)
ws05.cell(row=73, column=1, value='ETAPA 1').font = font(size=10, bold=True, color='854D0E')
c = ws05.cell(row=73, column=2, value='Pre-lanzamiento (deploy) — ver detalle hoja 05b')
c.font = font(size=10, bold=True, color='854D0E'); c.alignment = left()
# Por marca: pull from 05b
for i, m in enumerate(MARCAS):
    cell_ref = f"'05b · Etapa 1 deploy'!D{totals_per_marca[m]}"
    c = ws05.cell(row=73, column=3+i, value=f'={cell_ref}')
    c.number_format = MONEY
    c.font = font(size=10, color=GREEN_LINK, bold=True)
    c.alignment = right()
# Total portafolio (col K = 11)
c = ws05.cell(row=73, column=11, value=f'=SUM({get_column_letter(3)}73:{get_column_letter(10)}73)')
c.number_format = MONEY; c.font = font(size=10, bold=True, color='854D0E'); c.alignment = right()

# Row 74: Marketing Q2 → Etapa 2 (sin cambio en fórmulas, solo etiqueta)
ws05.cell(row=74, column=1, value='ETAPA 2').font = font(size=10, bold=True, color='854D0E')
c = ws05.cell(row=74, column=2, value='Comercialización (Marketing Q2 — 3 meses post-launch)')
c.font = font(size=10, bold=True, color='854D0E'); c.alignment = left()
# Las fórmulas de cada marca ya son correctas (=C63 etc.)

# Row 75: Renta → Etapa 3 (Contingencia 10%)
ws05.cell(row=75, column=1, value='ETAPA 3').font = font(size=10, bold=True, color='854D0E')
c = ws05.cell(row=75, column=2, value='Contingencia (10% sobre Etapa 1 + Etapa 2)')
c.font = font(size=10, bold=True, color='854D0E'); c.alignment = left()
# Fórmulas: 10% × (E1 + E2) por marca
for i, m in enumerate(MARCAS):
    col = get_column_letter(3+i)
    c = ws05.cell(row=75, column=3+i, value=f'=ROUND(({col}73+{col}74)*0.1,2)')
    c.number_format = MONEY
    c.font = font(size=10, color=GREEN_LINK, bold=True)
    c.alignment = right()
c = ws05.cell(row=75, column=11, value=f'=SUM({get_column_letter(3)}75:{get_column_letter(10)}75)')
c.number_format = MONEY; c.font = font(size=10, bold=True, color='854D0E'); c.alignment = right()

# Row 76 SUBTOTAL: cambiar etiqueta
c = ws05.cell(row=76, column=1, value='SUBTOTAL')
c.font = font(size=10, bold=True, color='854D0E')
c = ws05.cell(row=76, column=2, value='POR invertir al lanzamiento  (Etapas 1 + 2 + 3)')
c.font = font(size=11, bold=True, color='854D0E'); c.alignment = left()
# Las fórmulas =SUM(C73:C75) etc. siguen válidas — siguen sumando E1+E2+E3

# Row 79 nota — actualizar
ws05.merge_cells(start_row=79, start_column=1, end_row=79, end_column=11)
c = ws05.cell(row=79, column=1,
              value=('🎯  CAPITAL POR INVERTIR explicado a tus socios:  '
                     'Etapa 1 = lo que falta pagar para hacer DEPLOY (servicios técnicos, registros legales — ver hoja 05b detalle).  '
                     'Etapa 2 = los 3 meses de marketing post-lanzamiento (mayo-julio 2026).  '
                     'Etapa 3 = 10% de contingencia por imprevistos (si no se usa, queda en caja).  '
                     '⚠️ La renta del coworking ($90K Q2-Q4) NO está en Por invertir — la absorbe el holding iBisne como overhead.'))
c.font = font(size=9, italic=True, color='4B5563')
c.alignment = left(wrap=True)
c.fill = fill(LIGHT_GRAY)
ws05.row_dimensions[79].height = 50

# ============================================================================
# Reordenar hojas para que 05b vaya después de 05
# ============================================================================
order = ['00 · Dashboard', '01 · Portafolio', '02 · Avance', '03 · Canales',
         '04 · Reglas y plazos', '05 · Inversión iBisne', '05b · Etapa 1 deploy',
         '06 · P&L proyectado', '07 · Glosario', '08 · Bitácora']
wb._sheets = [wb[n] for n in order]
wb.active = 0

wb.save(PATH)
print(f'OK -> {PATH}')
print(f'Hoja 05b creada con 8 marcas × 10 conceptos = 80 filas editables')
print(f'IMPI pre-llenado en 5 marcas pendientes ($3,126.40 c/u = $15,632 total)')
print(f'Demás 75 filas en $0 — listas para que el usuario llene')
